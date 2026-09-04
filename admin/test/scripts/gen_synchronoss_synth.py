"""
Generate a synthetic Synchronoss/Verizon Cloud return for the synchronoss module tests.

Every value here is fabricated. A Synchronoss return is a provider legal return, not a
device extraction: there is no public image it appears on, it cannot be produced with a
throwaway account, and it is subscriber PII end to end. The quarantine container it
carries is the content that triggered the NCMEC CyberTip, so it cannot be sanitised or
shared in any form. A format-faithful synthetic return is therefore the only fixture this
module can have, and ground truth is known for every row it emits.

The shape is faithful to real returns, including the parts that have broken the parser:

  - all three DV access-log delivery shapes, which real returns mix by production rather
    than by date: monthly "Dv Access logs mdn ... .csv" files, a single "<LCID>.xlsx",
    and a "<LCID>_Dv_Access_Logs.xlsx" that renames the timestamp column to logtimestamp
    and writes Apache/CLF timestamps, one of them with a non-UTC offset
  - a workbook whose column titles are capitalised, which a file written for a person to
    read may well be, and which must still parse
  - a NUMBER in the querystring column of a lower-case-header workbook. It has to be in a
    lower-case file: with a header-case defect present the title-case columns read blank,
    so the numeric cell never reaches a string method and the two defects mask each other
  - a decoy workbook with the right extension and the wrong columns, which must be ignored
  - quarantine members named <container>_<sha256>.zip_file_<N>, including one whose
    filename hash deliberately disagrees with its bytes so the verification path is proven
  - extensionless media referenced only through a SMIL placeholder, and one filename
    present in two date folders, which is how media was mislinked before

Usage:
    python admin/test/scripts/gen_synchronoss_synth.py [output_dir]

Then zip the result and feed it to make_test_data.py to regenerate the committed fixture.
"""
import csv, json, os, io, shutil, sys
from datetime import datetime, timezone
from PIL import Image

BASE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "synth_return")
LCID = "SYNTH001LCIDHASH"
if os.path.exists(BASE):
    shutil.rmtree(BASE)

def mk(*parts):
    p = os.path.join(BASE, *parts)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    return p

def jpeg_bytes(color):
    b = io.BytesIO(); Image.new("RGB", (160, 100), color).save(b, "JPEG"); return b.getvalue()

def png_bytes(color):
    b = io.BytesIO(); Image.new("RGB", (160, 100), color).save(b, "PNG"); return b.getvalue()

def write_bytes(path, data):
    with open(path, "wb") as handle: handle.write(data)

def epoch_ms(iso):
    dt = datetime.strptime(iso, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
    return int(dt.timestamp() * 1000)

def mid(iso, h):  # Message ID = <epoch_ms>:<hash>; epoch prefix must == Date in ms
    return f"{epoch_ms(iso)}:{h}"

H1 = "a"*64; H2 = "b"*64; HDUP = "c"*64

# --- media files ----------------------------------------------------------
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-01", H1 + ".jpg"), jpeg_bytes((200,30,30)))
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-01", "recording000000.amr"), b"#!AMR\n" + b"\x00"*64)
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-01", H2 + ".jpg"), jpeg_bytes((30,140,60)))
# Extensionless real-media probe: a JPEG literally named "0" in the in/ folder
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-01", "0"), jpeg_bytes((240,200,20)))
# Sent media
write_bytes(mk(LCID, "messages/attachments/mms/out/2025-12-01", "image000000.jpg"), jpeg_bytes((40,60,200)))
# Cross-date same-name mislink probe: dup.jpg exists in TWO in/ date folders
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-01", "dup.jpg"), jpeg_bytes((150,0,150)))  # purple = Dec 1
write_bytes(mk(LCID, "messages/attachments/mms/in/2025-12-02", "dup.jpg"), jpeg_bytes((0,150,150)))   # teal   = Dec 2

# --- messages CSV day 1 ---------------------------------------------------
hdr = ["Date","Type","Direction","Attachments","Body","Sender","Recipients","Message ID"]
ACCT = "+12085550000"
def row(date,typ,dirn,att,body,sndr,rcpt,h): return [date,typ,dirn,att,body,sndr,rcpt,mid(date,h)]

day1 = [
 row("2025-12-01T00:13:17.000Z","sms","in","null.txt","Hey there 👋",     "+12085551111",ACCT,"d1"),
 row("2025-12-01T00:14:00.000Z","sms","out","null.txt","Reply emoji 😀",  "",            "+12085551111","d2"),
 row("2025-12-01T01:00:00.000Z","mms","in",f"smil;{H1}.jpg","",            "+12085552222",ACCT,"d3"),
 row("2025-12-01T01:05:00.000Z","mms","in","smil;recording000000.amr","", "+12085552222",ACCT,"d4"),
 # group MMS: many recipients -> must stack in report
 row("2025-12-01T02:00:00.000Z","mms","in",f"smil;{H2}.jpg","",            "+12085553333","+12085550000;+12085554444;+12085555555","d5"),
 # extensionless "0" real-media probe (current code SKIPS bare '0' token)
 row("2025-12-01T03:00:00.000Z","mms","in","smil;0","",                    "+12085552222",ACCT,"d6"),
 # sent MMS
 row("2025-12-01T04:00:00.000Z","mms","out","image000000.jpg","",          "",            "+12085551111","d7"),
 # calls
 row("2025-12-01T05:00:00.000Z","call","in","null.txt","",                 "+12085557777",ACCT,"d8"),
 row("2025-12-01T05:10:00.000Z","call","out","null.txt","",                "",            "+12085558888","d9"),
 # placeholder-only rows -> no media
 row("2025-12-01T06:00:00.000Z","mms","in","null.smi;0;1","",              "+12085552222",ACCT,"d10"),
 row("2025-12-01T06:10:00.000Z","mms","in","smil.xml;text000001.txt","",   "+12085552222",ACCT,"d11"),
 # mislink probe: message dated Dec 3 (no matching folder) references dup.jpg -> forces fallback
 row("2025-12-03T09:00:00.000Z","mms","in","smil;dup.jpg","",              "+12085559999",ACCT,"d12"),
 # quarantine probe: references a hash-named jpg with NO file on disk (flagged->removed)
 row("2025-12-01T07:00:00.000Z","mms","in",f"smil;{'9'*64}.jpg","",        "+12085552222",ACCT,"d14"),
]
with open(mk(LCID,"messages","20251201.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(hdr); w.writerows(day1)

# --- messages CSV day 2 ---------------------------------------------------
day2 = [
 row("2025-12-02T10:00:00.000Z","mms","in","smil;dup.jpg","",              "+12085551111",ACCT,"d13"),  # correct dup -> Dec2 teal
]
with open(mk(LCID,"messages","20251202.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(hdr); w.writerows(day2)

# --- VZMOBILE -------------------------------------------------------------
write_bytes(mk(LCID,"VZMOBILE/2025-12-01/My Samsung SM-F966U","backup000001.png"), png_bytes((90,90,90)))
write_bytes(mk(LCID,"VZMOBILE/2025-12-01/My Samsung SM-F966U","backup000002.png"), png_bytes((180,120,40)))
write_bytes(mk(LCID,"VZMOBILE/2025-12-01/My Samsung SM-F966U","noext_image"),     jpeg_bytes((20,180,180)))  # extensionless -> magic bytes

# --- contacts JSON --------------------------------------------------------
contacts = {"contacts":{"itemcount":3,"contact":[
 {"firstname":"Alice","lastname":"Active","source":"cloud","created":"Sat Jun  2 23:19:50 2018",
  "itemguid":"g1","incaseofemergency":False,"favorite":True,
  "tel":[{"type":"Mobile","indx":1,"number":"+1 714-487-1148","preference":0},
         {"type":"Home","indx":2,"number":"+1 208-555-0101","preference":0}]},
 {"firstname":"Bob","lastname":"Deleted","source":"cloud","created":"Sun Jan  6 10:00:00 2019",
  "deleted":"Wed May 22 12:59:18 2019","itemguid":"g2","incaseofemergency":False,"favorite":False,
  "tel":[{"type":"Mobile","indx":1,"number":"+1 208-555-0202","preference":0}]},
 {"firstname":"Carol","lastname":"NoPhone","source":"cloud","created":"Mon Mar  4 08:00:00 2019",
  "itemguid":"g3","incaseofemergency":True,"favorite":False,"tel":[]},
]}}
with open(mk("contacts_20251201.txt"),"w",encoding="utf-8") as f: json.dump(contacts,f,indent=1)

# --- DV access log --------------------------------------------------------
dv_hdr=["server_ts","remoteipaddress","clientidentifier","querystring","lcid"]
# Every DV row emitted, in column order [ts, remoteip, clientid, querystring, lcid],
# so the upload/sync ground truth below is counted from what was written rather
# than restated by hand and left to drift.
DV_ALL = []
dv=[
 # UPLOAD events (checksum present) - multi-IP quoted (user IP first, CDN rest)
 ["2025-12-01 01:00:05","174.21.5.9, 23.45.6.7, 23.45.6.8","SAMSUNG/SM-F966U",f"?checksum={'e'*64}&skipMissingFiles=true",LCID],
 ["2025-12-01 02:00:10","174.21.5.9","SAMSUNG/SM-F966U",f"?checksum={'f'*64}&skipMissingFiles=true",LCID],
 # SYNC events (no checksum)
 ["2025-12-01 02:05:00","174.21.5.9","SAMSUNG/SM-F966U","?conflictSolve=copyIfDifferent",LCID],
 ["2025-12-01 02:06:00","-","HF","?conflictSolve=copyIfDifferent",LCID],   # '-' IP, HF device
]
with open(mk(f"Dv Access logs mdn {LCID} December 2025.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(dv_hdr); w.writerows(dv)
DV_ALL += dv

# --- DV access log, monthly CSV with capitalised column titles ---------------
# The CSV twin of the title-case workbook below. This branch is identified by
# FILENAME, not by its columns, so a capitalised title row is accepted and then
# read blank unless the row keys are lower-cased too: right row count, every
# column empty, upload rows losing their checksum and refiling as sync. Machine
# written files are likelier lower case, but it is the same failure class, and
# the guard costs one file.
dv_title_hdr = ["Server_TS","RemoteIPAddress","ClientIdentifier","QueryString","LCID"]
dv_title = [
 ["2026-02-02 03:00:00","174.21.5.9, 23.45.6.7","SAMSUNG/SM-F966U",f"?checksum={'a7'*32}&skipMissingFiles=true",LCID],
 ["2026-02-02 03:05:00","174.21.5.9","SAMSUNG/SM-F966U",f"?checksum={'b8'*32}&skipMissingFiles=true",LCID],
 ["2026-02-02 03:10:00","174.21.5.9","SAMSUNG/SM-F966U","?conflictSolve=copyIfDifferent",LCID],
]
with open(mk(f"Dv Access logs mdn {LCID} February 2026.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(dv_title_hdr); w.writerows(dv_title)
DV_ALL += dv_title

# --- quarantined ("CyberTip") container ------------------------------------
# Real returns deliver <LCID>-<container>-quarantined.zip, which extracts to a
# doubly-nested folder of files named <container>_<sha256>.zip_file_<N>. Despite
# the name these are NOT split-archive parts: each is a complete standalone media
# file whose SHA-256 is its own filename. Mirror that exactly, including the
# double nesting, so the glob and the name regex are both exercised.
import hashlib
from openpyxl import Workbook

CONTAINER = "770d2225305340cea852d92626230d0b"
QDIR = os.path.join(f"{LCID}-{CONTAINER}-quarantined", f"{LCID}-{CONTAINER}-quarantined")

def quarantine(seq, data, force_hash=None):
    """Write one quarantine member. force_hash fakes a wrong name-hash on purpose."""
    real = hashlib.sha256(data).hexdigest()
    name = f"{CONTAINER}_{force_hash or real}.zip_file_{seq}"
    write_bytes(mk(QDIR, name), data)
    return real

# q1/q2 correlate to DV upload events; q3 has no upload event; q4's filename hash
# deliberately disagrees with its bytes, to prove the verification actually fires.
Q1 = quarantine(1, jpeg_bytes((10, 10, 10)))
Q2 = quarantine(2, png_bytes((20, 90, 20)))
Q3 = quarantine(3, jpeg_bytes((90, 20, 90)))
quarantine(4, jpeg_bytes((5, 5, 200)), force_hash="d" * 64)

# --- DV access log, workbook form (2026-format returns) --------------------
# Same five columns, but delivered as a single <LCID>.xlsx whose filename carries
# no "DV" marker — so the parser must accept it on its headers, not its name.
wb = Workbook(); ws = wb.active
ws.append(dv_hdr)
dv_wb1 = [
    ["2026-01-04 08:15:00", "174.21.5.9, 23.45.6.7", "SAMSUNG/SM-F966U", f"?checksum={Q1}&skipMissingFiles=true", LCID],
    ["2026-01-04 09:20:00", "174.21.5.9", "SAMSUNG/SM-F966U", f"?checksum={Q2}&skipMissingFiles=true", LCID],
    ["2026-01-05 11:00:00", "10.0.0.5", "SAMSUNG/SM-F966U", f"?checksum={'9'*64}&skipMissingFiles=true", LCID],
    ["2026-01-05 11:05:00", "10.0.0.5", "SAMSUNG/SM-F966U", "?conflictSolve=copyIfDifferent", LCID],
]
for r in dv_wb1:
    ws.append(r)
DV_ALL += dv_wb1
wb.save(mk(f"{LCID}.xlsx"))

# --- DV access log, THIRD variant seen in the wild ---------------------------
# Underscored filename, "logtimestamp" instead of "server_ts", and Apache/CLF
# timestamps with an explicit offset instead of "YYYY-MM-DD HH:MM:SS".
wb3 = Workbook(); ws3 = wb3.active
ws3.append(["logtimestamp", "remoteipaddress", "clientidentifier", "querystring", "lcid"])
dv_wb3 = [
    ["[02/Jun/2026:14:23:11 +0000]", "174.21.5.9, 23.45.6.7", "SAMSUNG/SM-F966U", f"?checksum={'a1'*32}&skipMissingFiles=true", LCID],
    ["[02/Jun/2026:14:23:12 -0600]", "174.21.5.9", "SAMSUNG/SM-F966U", f"?checksum={'b2'*32}&skipMissingFiles=true", LCID],
    ["[02/Jun/2026:14:25:00 +0000]", "174.21.5.9", "SAMSUNG/SM-F966U", "?conflictSolve=copyIfDifferent", LCID],
    # A NUMBER in the querystring column of a LOWER-CASE-header workbook. It has to be
    # here rather than only in the title-case file: with the header-case bug present the
    # title-case columns read blank, so the numeric cell never reaches a string method
    # and the two defects mask each other. This row reaches the code either way.
    ["[02/Jun/2026:14:26:00 +0000]", "174.21.5.9", "SAMSUNG/SM-F966U", 20260602, LCID],
]
for r in dv_wb3:
    ws3.append(r)
DV_ALL += dv_wb3
wb3.save(mk(f"{LCID}_Dv_Access_Logs.xlsx"))

# --- DV access log, FOURTH shape: capitalised column titles -------------------
# A workbook written for a person to read may title-case its columns. Header
# matching is case-insensitive, so this file is accepted; if the row keys were not
# lower-cased too it would then read back empty - accepted and silent, which is the
# failure this module exists to prevent. One row also carries a NUMBER in the
# querystring column: a workbook cell need not be text, and a consumer calling a
# string method on the raw cell would fail the whole artifact rather than one row.
wb4 = Workbook(); ws4 = wb4.active
ws4.append(["Server_TS", "RemoteIPAddress", "ClientIdentifier", "QueryString", "LCID"])
dv_wb4 = [
    ["2026-03-11 07:00:00", "174.21.5.9, 23.45.6.7", "SAMSUNG/SM-F966U", f"?checksum={'c3'*32}&skipMissingFiles=true", LCID],
    ["2026-03-11 07:05:00", "174.21.5.9", "SAMSUNG/SM-F966U", f"?checksum={'d4'*32}&skipMissingFiles=true", LCID],
    ["2026-03-11 07:10:00", "174.21.5.9", "SAMSUNG/SM-F966U", "?conflictSolve=copyIfDifferent", LCID],
    ["2026-03-11 07:15:00", "174.21.5.9", "SAMSUNG/SM-F966U", 20260311, LCID],
]
for r in dv_wb4:
    ws4.append(r)
DV_ALL += dv_wb4
wb4.save(mk(f"{LCID}_TitleCase_Dv_Access_Logs.xlsx"))

# Decoy workbook: right extension, wrong columns. Must be ignored, not parsed.
wb2 = Workbook(); ws2 = wb2.active
ws2.append(["case", "examiner", "notes"]); ws2.append(["SYNTH", "nobody", "not a DV log"])
wb2.save(mk("case_notes.xlsx"))

# --- ground truth ------------------------------------------------------------
# Stated independently of the parser, then checked against what was actually
# emitted. A known-values line that is only printed drifts silently as probes are
# added -- this file's did -- and once it has drifted the fixture is no longer
# proving anything beyond "whatever the parser said". Asserting closes that.
EXPECTED = {
    "sms_mms_rows":   12,   # 11 in 20251201.csv (9 mms + 2 sms) + 1 in 20251202.csv
    "call_rows":       2,
    "contacts":        3,   # 1 active, 1 deleted, 1 with no phone number
    "dv_uploads":     11,   # rows carrying a checksum in the querystring
    "dv_sync":         8,   # rows without one
    "quarantined":     4,   # 2 correlated to a DV upload, 1 not, 1 wrong name-hash
    "vzmobile":        3,   # 2 png + 1 extensionless (typed by magic bytes)
    "mms_media_files": 7,   # in/ and out/ media, incl. "0" and both dup.jpg copies
}

msg_rows = day1 + day2
actual = {
    "sms_mms_rows":   sum(1 for r in msg_rows if r[1] in ("sms", "mms")),
    "call_rows":      sum(1 for r in msg_rows if r[1] == "call"),
    "contacts":       len(contacts["contacts"]["contact"]),
    "dv_uploads":     sum(1 for r in DV_ALL if "checksum=" in str(r[3])),
    "dv_sync":        sum(1 for r in DV_ALL if "checksum=" not in str(r[3])),
    "quarantined":    len([n for n in os.listdir(mk(QDIR, ".")) if "zip_file_" in n]),
    "vzmobile":       3,
    "mms_media_files": sum(
        len(files)
        for root, _, files in os.walk(os.path.join(BASE, LCID, "messages", "attachments", "mms"))
    ),
}

mismatches = {k: (EXPECTED[k], actual[k]) for k in EXPECTED if EXPECTED[k] != actual[k]}
if mismatches:
    raise SystemExit(
        "generator ground truth is stale, fix EXPECTED or the data: "
        + ", ".join(f"{k}: declared {d}, emitted {a}" for k, (d, a) in mismatches.items()))

print("Synthetic return generated at:", BASE)
print(f"Ground truth (asserted): {actual['sms_mms_rows']} SMS/MMS msgs, "
      f"{actual['call_rows']} calls, {actual['contacts']} contacts"
      " (1 deleted, 1 no-phone),")
print(f"  DV: {actual['dv_uploads']} uploads + {actual['dv_sync']} sync across 5 files")
print("      [CSV 2+2, title-case CSV 2+1, <LCID>.xlsx 3+1,")
print("       _Dv_Access_Logs.xlsx 2+2, _TitleCase_Dv_Access_Logs.xlsx 2+2]")
print("      title-case CSV and workbook both guard the header-case defect, one per reader;")
print("      the two lower-case files carry the NUMERIC querystring cells, which a")
print("      title-case-only placement would mask behind the header-case defect;")
print("      _Dv_Access_Logs.xlsx uses logtimestamp + Apache/CLF stamps incl. a -0600 offset,")
print(f"  VZMOBILE: {actual['vzmobile']} files (2 png + 1 extensionless),")
print(f"  Quarantined: {actual['quarantined']} members - 2 correlated to a DV upload, 1 uncorrelated,")
print("               1 with a deliberately wrong filename hash (must report NOT verified),")
print(f"  MMS media: {actual['mms_media_files']} files,")
print("  probes: '0' extensionless real media (referenced as bare '0'), dup.jpg in two date")
print("          folders (order-independence probe), case_notes.xlsx decoy workbook")
print("          (must NOT be read as a DV log)")
