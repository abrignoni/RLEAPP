__artifacts_v2__ = {
    "icloudReturnsLogs": {
        "name": "iCloud - Logs",
        "description": "iCloud access logs from an iCloud law enforcement return (xlsx).",
        "author": "@AlexisBrignoni",
        "creation_date": "2021-08-18",
        "last_update_date": "2026-06-27",
        "requirements": "openpyxl",
        "category": "iCloud Returns",
        "notes": "",
        "paths": ('*/LOG/*_iCloudLogs.xlsx',),
        "output_types": "standard",
        "artifact_icon": "list",
    }
}

import os
from datetime import date, time

from openpyxl import load_workbook

from scripts.ilapfuncs import artifact_processor
from scripts.lavafuncs import sanitize_sql_name

_RESERVED = {'from', 'to', 'order', 'group', 'where', 'select', 'index', 'join', 'references',
             'check', 'default', 'add', 'table', 'column', 'create', 'insert', 'update', 'delete',
             'drop', 'values', 'set', 'primary', 'key', 'unique', 'foreign', 'constraint', 'having',
             'distinct', 'union', 'using'}


def _safe_headers(cells):
    seen, out = {}, []
    for i, cell in enumerate(cells):
        name = str(cell).strip() if cell not in (None, '') else f'Column {i + 1}'
        if sanitize_sql_name(name) in _RESERVED:
            name = f'{name} Value'
        key = name.lower()
        seen[key] = seen.get(key, -1) + 1
        if seen[key]:
            name = f'{name} ({seen[key]})'
        out.append(name)
    return out


def _parse_icloud_xlsx(file_found, header_row):
    workbook = load_workbook(file_found, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    headers, rows = [], []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < header_row:
            continue
        cells = [str(c) if isinstance(c, (date, time)) else ('' if c is None else c) for c in row]
        if i == header_row:
            headers = _safe_headers(cells)
        else:
            rows.append(cells)
    workbook.close()
    return headers, rows


@artifact_processor
def icloudReturnsLogs(context):
    headers, data_list, source_path = [], [], ''
    for file_found in context.get_files_found():
        file_found = str(file_found)
        basename = os.path.basename(file_found)
        if basename.startswith('~') or basename.startswith('.'):
            continue
        source_path = file_found
        sheet_headers, rows = _parse_icloud_xlsx(file_found, 6)
        if sheet_headers:
            headers = sheet_headers
            data_list.extend(rows)

    return tuple(headers), data_list, context.get_relative_path(source_path)
