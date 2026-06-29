_REGEX_NOTE = ("Scans every cell of the return workbook with regexes; the account token comes "
               "from the file name.")
_SECTION_NOTE = ("Walks a named header section of the return workbook (reading down/right from the "
                 "header cell) to surface the account data Cash App reports as labelled tables. "
                 "The original module parsed this data into an AccountInfo object but never "
                 "reported it; these artifacts surface it. The account token comes from the file "
                 "name.")


def _meta(name, icon, notes, html_columns=None):
    meta = {"name": f"CashApp - {name}",
            "description": f"{name} extracted from a Cash App law enforcement return "
                           f"(*-for-subject-SQ_CASH-*.xlsx).",
            "author": "Shawn Ramsey", "creation_date": "2024-02-02",
            "last_update_date": "2026-06-28", "requirements": "openpyxl",
            "category": "Cash App Returns", "notes": notes,
            "paths": ('*/*-for-subject-SQ_CASH-*.xlsx',), "output_types": "standard",
            "artifact_icon": icon}
    if html_columns:
        meta["html_columns"] = html_columns
    return meta


__artifacts_v2__ = {
    "cashappEmails": _meta("Emails", "mail", _REGEX_NOTE),
    "cashappIPv4": _meta("IPv4", "globe", _REGEX_NOTE),
    "cashappIPv6": _meta("IPv6", "globe", _REGEX_NOTE),
    "cashappPhoneNumbers": _meta("Phone Numbers", "phone", _REGEX_NOTE),
    "cashappDisplayNames": _meta("Display Name History", "user", _SECTION_NOTE),
    "cashappPaymentCards": _meta("Payment Source Cards", "credit-card", _SECTION_NOTE),
    "cashappPaymentBankAccounts": _meta("Payment Source Bank Accounts", "dollar-sign",
                                        _SECTION_NOTE),
    "cashappIssuedVirtualCards": _meta("Issued Virtual Cards", "credit-card", _SECTION_NOTE),
    "cashappIssuedPhysicalCards": _meta("Issued Physical Cards", "credit-card", _SECTION_NOTE),
    "cashappIssuedBankAccounts": _meta("Issued Bank Accounts", "dollar-sign", _SECTION_NOTE),
}

import os
import re
from datetime import datetime, timezone

from openpyxl import load_workbook

from scripts.ilapfuncs import artifact_processor, convert_unix_ts_to_utc, ipgen

_ACCOUNT_TOKEN_SUBSTR = "-for-subject-SQ_CASH-"

_email_re = re.compile(r"^[a-z0-9!#$%&'*+/=?^_`{|}~-]+(?:\.[a-z0-9!#$%&'*+/=?^_`{|}~-]+)*@"
                       r"(?:[a-z0-9](?:[a-z0-9-]*[a-z0-9])?\.)+[a-z0-9](?:[a-z0-9-]*[a-z0-9])?$")
_ipv4_re = re.compile(r"^(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.(?:25[0-5]|2[0-4][0-9]|"
                      r"[01]?[0-9][0-9]?)\.(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\."
                      r"(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$")
_ipv6_re = re.compile(
    r"^(([0-9a-fA-F]{1,4}:){7,7}[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,7}:|"
    r"([0-9a-fA-F]{1,4}:){1,6}:[0-9a-fA-F]{1,4}|([0-9a-fA-F]{1,4}:){1,5}(:[0-9a-fA-F]{1,4}){1,2}|"
    r"([0-9a-fA-F]{1,4}:){1,4}(:[0-9a-fA-F]{1,4}){1,3}|([0-9a-fA-F]{1,4}:){1,3}"
    r"(:[0-9a-fA-F]{1,4}){1,4}|([0-9a-fA-F]{1,4}:){1,2}(:[0-9a-fA-F]{1,4}){1,5}|"
    r"[0-9a-fA-F]{1,4}:((:[0-9a-fA-F]{1,4}){1,6})|:((:[0-9a-fA-F]{1,4}){1,7}|:)|"
    r"fe80:(:[0-9a-fA-F]{0,4}){0,4}%[0-9a-zA-Z]{1,}|::(ffff(:0{1,4}){0,1}:){0,1}"
    r"((25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3,3}(25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])|"
    r"([0-9a-fA-F]{1,4}:){1,4}:((25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9])\.){3,3}"
    r"(25[0-5]|(2[0-4]|1{0,1}[0-9]){0,1}[0-9]))$")
_phone_re = re.compile(r"^\+?[1-9][0-9]{7,14}$")


def _token(file_found):
    idx = file_found.find(_ACCOUNT_TOKEN_SUBSTR)
    if idx == -1:
        return os.path.basename(file_found)
    token = file_found[idx + len(_ACCOUNT_TOKEN_SUBSTR):]
    return token[:-5] if token.endswith('.xlsx') else token


def _to_utc(value):
    if value is None or value == '':
        return ''
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value.astimezone(timezone.utc)
    text = str(value).strip()
    if not text:
        return ''
    if text.isdigit():
        return convert_unix_ts_to_utc(int(text))
    try:
        dt = datetime.fromisoformat(text.replace('Z', '+00:00'))
        return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt.astimezone(timezone.utc)
    except ValueError:
        return value


def _scan(file_found, token):
    emails, ipv4s, ipv6s, phones = [], [], [], []
    workbook = load_workbook(file_found, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    value = str(sheet.cell(row=i, column=j).value)
                    if _email_re.match(value):
                        date = sheet.cell(row=i, column=j - 1).value if j > 1 else None
                        emails.append((token, _to_utc(date), value))
                    if _ipv4_re.match(value):
                        ipv4s.append((token, value))
                    if _ipv6_re.match(value):
                        ipv6s.append((token, value))
                    if _phone_re.match(value):
                        above = str(sheet.cell(row=i - 1, column=j).value) if i > 1 else ''
                        if "Full SSN" not in above:
                            date = sheet.cell(row=i, column=j - 1).value if j > 1 else None
                            phones.append((token, _to_utc(date), value))
    finally:
        workbook.close()
    return emails, ipv4s, ipv6s, phones


def _cell(value):
    """Render a workbook cell as clean text, trimming the float '.0' that openpyxl puts on
    integer-valued numeric cells (card/bank/routing numbers, zip codes)."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _walk_display_names(sheet, i, j, token, out):
    # Values run down the same column from the row below the header until a blank cell.
    a = 1
    while sheet.cell(row=i + a, column=j).value is not None:
        out.append((token, _cell(sheet.cell(row=i + a, column=j).value)))
        a += 1


def _walk_payment_source(sheet, i, j, token, cards, banks):
    # Card rows (number, brand, zipcode) start 2 rows below the header (row i+1 is the sub-header).
    a = 2
    while sheet.cell(row=i + a, column=j).value is not None:
        cards.append((token,
                      _cell(sheet.cell(row=i + a, column=j).value),
                      _cell(sheet.cell(row=i + a, column=j + 1).value),
                      _cell(sheet.cell(row=i + a, column=j + 2).value)))
        a += 1
    # One blank row separates the cards from the "Bank Account Number" sub-section.
    a += 1
    if str(sheet.cell(row=i + a, column=j).value) == 'Bank Account Number':
        a += 1
        while sheet.cell(row=i + a, column=j).value is not None:
            banks.append((token,
                          _cell(sheet.cell(row=i + a, column=j).value),
                          _cell(sheet.cell(row=i + a, column=j + 1).value)))
            a += 1


def _walk_issued(sheet, i, j, token, virtual, physical, banks):
    # Virtual card rows (number, issued_date) start 2 rows below the header.
    a = 2
    while sheet.cell(row=i + a, column=j).value is not None:
        virtual.append((token,
                        _cell(sheet.cell(row=i + a, column=j).value),
                        _to_utc(sheet.cell(row=i + a, column=j + 1).value)))
        a += 1
    # Blank row, then the optional "Physical Card Number" sub-section (number, issued_date, address).
    a += 1
    if str(sheet.cell(row=i + a, column=j).value) == 'Physical Card Number':
        a += 1
        while sheet.cell(row=i + a, column=j).value is not None:
            physical.append((token,
                             _cell(sheet.cell(row=i + a, column=j).value),
                             _to_utc(sheet.cell(row=i + a, column=j + 1).value),
                             _cell(sheet.cell(row=i + a, column=j + 2).value)))
            a += 1
    # Blank row, then the optional "Bank Account Number" sub-section (account, routing). This
    # advance runs whether or not a physical-card section was present (matching the original walk).
    a += 1
    if str(sheet.cell(row=i + a, column=j).value) == 'Bank Account Number':
        a += 1
        while sheet.cell(row=i + a, column=j).value is not None:
            banks.append((token,
                          _cell(sheet.cell(row=i + a, column=j).value),
                          _cell(sheet.cell(row=i + a, column=j + 1).value)))
            a += 1


def _walk_sections(file_found, token):
    sections = {'display_names': [], 'cards': [], 'banks': [],
                'virtual_cards': [], 'physical_cards': [], 'issued_banks': []}
    workbook = load_workbook(file_found, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for i in range(1, sheet.max_row + 1):
                for j in range(1, sheet.max_column + 1):
                    header = str(sheet.cell(row=i, column=j).value)
                    if header == 'Display Name History':
                        _walk_display_names(sheet, i, j, token, sections['display_names'])
                    elif header == 'Payment Source History':
                        _walk_payment_source(sheet, i, j, token, sections['cards'],
                                             sections['banks'])
                    elif header == 'Issued Instruments':
                        _walk_issued(sheet, i, j, token, sections['virtual_cards'],
                                     sections['physical_cards'], sections['issued_banks'])
    finally:
        workbook.close()
    return sections


def _xlsx_files(context):
    for file_found in context.get_files_found():
        file_found = str(file_found)
        if file_found.endswith('.xlsx') and not os.path.basename(file_found).startswith('.'):
            yield file_found


@artifact_processor
def cashappEmails(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        emails, _, _, _ = _scan(file_found, _token(file_found))
        data_list.extend(emails)
    data_headers = ('Account Token', ('UTC Date Time', 'datetime'), 'Email')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappIPv4(context):
    data_list, ip_list, source_path = [], [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        token = _token(file_found)
        _, ipv4s, _, _ = _scan(file_found, token)
        data_list.extend(ipv4s)
        for _, ipv4 in ipv4s:
            ip_list.append((ipv4, 'CashApp', f'CashApp IPv4 Return - {token}', '', None))
    if ip_list:
        ipgen(context.get_report_folder(), ip_list)
    data_headers = ('Account Token', 'IPv4')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappIPv6(context):
    data_list, ip_list, source_path = [], [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        token = _token(file_found)
        _, _, ipv6s, _ = _scan(file_found, token)
        data_list.extend(ipv6s)
        for _, ipv6 in ipv6s:
            ip_list.append((ipv6, 'CashApp', f'CashApp IPv6 Return - {token}', '', None))
    if ip_list:
        ipgen(context.get_report_folder(), ip_list)
    data_headers = ('Account Token', 'IPv6')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappPhoneNumbers(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        _, _, _, phones = _scan(file_found, _token(file_found))
        data_list.extend(phones)
    data_headers = ('Account Token', ('UTC Date Time', 'datetime'), ('Phone Numbers', 'phonenumber'))
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappDisplayNames(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['display_names'])
    data_headers = ('Account Token', 'Display Name')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappPaymentCards(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['cards'])
    data_headers = ('Account Token', 'Card Number', 'Card Brand', 'Zip Code')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappPaymentBankAccounts(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['banks'])
    data_headers = ('Account Token', 'Bank Account Number', 'Routing Number')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappIssuedVirtualCards(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['virtual_cards'])
    data_headers = ('Account Token', 'Virtual Card Number', ('Issued Date', 'datetime'))
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappIssuedPhysicalCards(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['physical_cards'])
    data_headers = ('Account Token', 'Physical Card Number', ('Issued Date', 'datetime'), 'Address')
    return data_headers, data_list, context.get_relative_path(source_path)


@artifact_processor
def cashappIssuedBankAccounts(context):
    data_list, source_path = [], ''
    for file_found in _xlsx_files(context):
        source_path = file_found
        data_list.extend(_walk_sections(file_found, _token(file_found))['issued_banks'])
    data_headers = ('Account Token', 'Bank Account Number', 'Routing Number')
    return data_headers, data_list, context.get_relative_path(source_path)
